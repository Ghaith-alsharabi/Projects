from flask import render_template, request, redirect, url_for, current_app
from sport.renameCsv import bp, UPLOAD_FOLDER
from .rename import RenameForm, allowed_file
import os
from openpyxl import load_workbook

APP_ROOT = os.path.dirname(os.path.abspath(__file__))
SHEET_FOLDER = os.path.join(APP_ROOT, 'static/uploads')
current_app.config['SHEETS_FOLDER'] = SHEET_FOLDER

@bp.route("/dashboard_", methods=["GET", "POST"])
def renameCsv():
    form = RenameForm()
    if request.method == 'POST':
        print(request.files['file'])
        print(form.name.data)
        file = request.files['file']
        if not form.name.data == '':
            file.filename = form.name.data
        if form.validate_on_submit():
            file.save(os.path.join(current_app.config['SHEETS_FOLDER'], file.filename))
            return "form submitted " + file.filename
    return render_template("renameCsv/index.jinja2", form=form)

@bp.route("/overview", methods=["GET", "POST"])
def get_all_uploaded_csvs():
    targetDir = os.path.join(current_app.config['SHEETS_FOLDER'])
    files = os.listdir(targetDir)
    for f in files:
        print(f)
    return render_template("renameCsv/uploaded_csvs.jinja2", files=files)

@bp.route("/edit/<filename>", methods=["GET", "POST"])
def edit_csv(filename):
    file_path = os.path.join(current_app.config['SHEETS_FOLDER']) + "/" + filename
    sheet_names = load_workbook(file_path, read_only=True).sheetnames
    sheets = [str(sheet) for sheet in sheet_names]
    return render_template("renameCsv/edit_csv.jinja2", filename=filename, sheets=sheets)

